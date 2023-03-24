import openpyxl
import os
from prof import Prof
import datetime
import time

file_name = 'Classeur1.xlsx'
feuille_name = 'Absence Prof'
Nombre_Prof = 150

class Exel :
    def __init__(self):
        self.readFile()
        self.retake = True

    def readFile(self):
        # Ouvrir le fichier Excel
        self.wb = openpyxl.load_workbook(file_name)
        # Sélectionner une feuille spécifique
        self.feuille = self.wb[feuille_name]
        # Trouver le nombre de ligne de la feuille de calcul
        self.nombre_Prof = Nombre_Prof
        self.date_save = datetime.datetime.fromtimestamp(os.path.getmtime(file_name))
        self.modified_date = datetime.datetime.fromtimestamp(os.path.getmtime(file_name))


    def reload(self):

        self.date_save = self.modified_date

        self.modified_date = datetime.datetime.fromtimestamp(os.path.getmtime(file_name))

        if self.modified_date != self.date_save:

                self.date_save = self.modified_date
                self.retake = False
                return self.retake
                print('changement')



    def Data(self):
        self.readFile()
        print(self.nombre_Prof)

        for i in range(1, self.nombre_Prof ):
            i = i+1

            Input_Name = str(self.feuille.cell(row=i, column=1).value)
            #return self.Input_Name
            Input_Surname = str(self.feuille.cell(row=i, column=2).value)
            #return self.Input_Surname

            #Data_name = self.Input_Name
            Prof.name = Input_Name
            Prof.surname = Input_Surname

            # vidé la note (liste prof)

            prof_com = Prof(name=Input_Name,surname=Input_Surname)

            prof_com.Data_Absent()




    def Auto_run(self):
        while True:
            print("")
            print("reload")
            print("")
            self.Data()

            while self.retake:
                self.reload()
                time.sleep(1)

            time.sleep(1)
            self.retake = True




""" def convert(date_time):
        format = '%b %d %Y %I:%M%p'  # The format
        datetime_str = datetime.datetime.strptime(date_time, format)

        return datetime_str

instancier ou faire une referance

z = WWW 
apatchi 
endigs 

sinon Prenom.nom    .TTinfo.be"""


